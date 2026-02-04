"""
Export services for generating PDF, Excel, CSV files.
"""

import io
import csv
import json
from datetime import datetime
from typing import Dict, Any, List, Optional
from django.http import HttpResponse
from django.utils import timezone


class BaseExporter:
    """Base class for exporters"""
    
    def __init__(self, data: Dict[str, Any], title: str = 'Report'):
        self.data = data
        self.title = title
        self.generated_at = timezone.now()
    
    def export(self) -> bytes:
        raise NotImplementedError


class CSVExporter(BaseExporter):
    """Export data to CSV format"""
    
    def export(self, data_key: Optional[str] = None) -> bytes:
        output = io.StringIO()
        
        # Get data to export
        if data_key and data_key in self.data:
            export_data = self.data[data_key]
        elif isinstance(self.data, list):
            export_data = self.data
        else:
            # Flatten dict to list
            export_data = [self.data]
        
        if not export_data:
            return b''
        
        # Write CSV
        if isinstance(export_data, list) and len(export_data) > 0:
            if isinstance(export_data[0], dict):
                writer = csv.DictWriter(output, fieldnames=export_data[0].keys())
                writer.writeheader()
                writer.writerows(export_data)
            else:
                writer = csv.writer(output)
                for row in export_data:
                    writer.writerow([row] if not isinstance(row, (list, tuple)) else row)
        
        return output.getvalue().encode('utf-8-sig')
    
    def get_response(self, filename: str = 'report.csv') -> HttpResponse:
        content = self.export()
        response = HttpResponse(content, content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ExcelExporter(BaseExporter):
    """Export data to Excel format using openpyxl"""
    
    def export(self) -> bytes:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl kutubxonasi o'rnatilmagan. pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = self.title[:31]  # Excel sheet name limit
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title row
        ws.merge_cells('A1:F1')
        ws['A1'] = self.title
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Generated at
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Yaratilgan: {self.generated_at.strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        current_row = 4
        
        # Process data sections
        for section_name, section_data in self.data.items():
            if isinstance(section_data, dict):
                # Section header
                ws.cell(row=current_row, column=1, value=section_name.replace('_', ' ').title())
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
                current_row += 1
                
                # Key-value pairs
                for key, value in section_data.items():
                    ws.cell(row=current_row, column=1, value=key.replace('_', ' ').title())
                    ws.cell(row=current_row, column=2, value=str(value) if not isinstance(value, (int, float)) else value)
                    current_row += 1
                
                current_row += 1
                
            elif isinstance(section_data, list) and len(section_data) > 0:
                # Section header
                ws.cell(row=current_row, column=1, value=section_name.replace('_', ' ').title())
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
                current_row += 1
                
                if isinstance(section_data[0], dict):
                    # Table headers
                    headers = list(section_data[0].keys())
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=current_row, column=col, value=header.replace('_', ' ').title())
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border
                    current_row += 1
                    
                    # Table data
                    for item in section_data:
                        for col, key in enumerate(headers, 1):
                            value = item.get(key, '')
                            cell = ws.cell(row=current_row, column=col, value=value)
                            cell.border = thin_border
                        current_row += 1
                
                current_row += 1
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def get_response(self, filename: str = 'report.xlsx') -> HttpResponse:
        content = self.export()
        response = HttpResponse(
            content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class PDFExporter(BaseExporter):
    """Export data to PDF format using reportlab"""
    
    def export(self) -> bytes:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch, cm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            raise ImportError("reportlab kutubxonasi o'rnatilmagan. pip install reportlab")
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceBefore=20,
            spaceAfter=10
        )
        normal_style = styles['Normal']
        
        story = []
        
        # Title
        story.append(Paragraph(self.title, title_style))
        story.append(Paragraph(
            f"Yaratilgan: {self.generated_at.strftime('%Y-%m-%d %H:%M')}",
            normal_style
        ))
        story.append(Spacer(1, 20))
        
        # Process data sections
        for section_name, section_data in self.data.items():
            story.append(Paragraph(section_name.replace('_', ' ').title(), heading_style))
            
            if isinstance(section_data, dict):
                # Key-value table
                table_data = [
                    [key.replace('_', ' ').title(), str(value)]
                    for key, value in section_data.items()
                    if not isinstance(value, (dict, list))
                ]
                if table_data:
                    t = Table(table_data, colWidths=[200, 250])
                    t.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTSIZE', (0, 0), (-1, -1), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                        ('TOPPADDING', (0, 0), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ]))
                    story.append(t)
                    story.append(Spacer(1, 10))
                    
            elif isinstance(section_data, list) and len(section_data) > 0:
                if isinstance(section_data[0], dict):
                    # Table
                    headers = list(section_data[0].keys())
                    table_data = [
                        [h.replace('_', ' ').title() for h in headers]
                    ]
                    for item in section_data[:20]:  # Limit rows
                        row = [str(item.get(h, ''))[:50] for h in headers]  # Limit cell length
                        table_data.append(row)
                    
                    col_width = min(450 / len(headers), 150)
                    t = Table(table_data, colWidths=[col_width] * len(headers))
                    t.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('FONTSIZE', (0, 1), (-1, -1), 9),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                        ('TOPPADDING', (0, 0), (-1, -1), 6),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    ]))
                    story.append(t)
            
            story.append(Spacer(1, 15))
        
        doc.build(story)
        output.seek(0)
        return output.getvalue()
    
    def get_response(self, filename: str = 'report.pdf') -> HttpResponse:
        content = self.export()
        response = HttpResponse(content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class JSONExporter(BaseExporter):
    """Export data to JSON format"""
    
    def export(self) -> bytes:
        export_data = {
            'title': self.title,
            'generated_at': self.generated_at.isoformat(),
            'data': self.data
        }
        return json.dumps(export_data, ensure_ascii=False, indent=2, default=str).encode('utf-8')
    
    def get_response(self, filename: str = 'report.json') -> HttpResponse:
        content = self.export()
        response = HttpResponse(content, content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


def get_exporter(format: str, data: Dict[str, Any], title: str = 'Report') -> BaseExporter:
    """Factory function to get appropriate exporter"""
    exporters = {
        'csv': CSVExporter,
        'excel': ExcelExporter,
        'xlsx': ExcelExporter,
        'pdf': PDFExporter,
        'json': JSONExporter,
    }
    
    exporter_class = exporters.get(format.lower())
    if not exporter_class:
        raise ValueError(f"Noma'lum format: {format}. Mavjud formatlar: {', '.join(exporters.keys())}")
    
    return exporter_class(data, title)
